import pandas as pd
import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import filedialog, messagebox

# user selects one or more files to analyze
# filename must be in the following format:
# {flowrate}ml_m {input temp target}C batt{batery id#} disp{disposable id#}
# e.g. 92ml_m 10C battA4 disp6.xlsx
# alternatively the text can be ommitted
# e.g. 92 10 A4 6.xlsx
# user selects an excel output file to append data summary
# output file can contain previous data or not
# data calculations are performed and appended to input and output files


def main():

    # select input file(s)
    filetypes = [("excel files", ".xlsx"), ("all files", ".*")]
    filenames = ""
    filenames = filedialog.askopenfilenames(
        initialdir=os.getcwd(), title="Select file input file(s)", filetypes=filetypes
    )
    if not filenames:
        sys.exit()

    # select file to append data
    # titles will be added to sheet "Data Summary"
    # if "Data Summary" sheet does not exit, one will be created
    filename_out = filedialog.askopenfilename(
        title="Select output file", filetypes=filetypes
    )
    if not filename_out:
        sys.exit()
    try:
        wb = load_workbook(filename_out)
    except PermissionError:
        show_error(f"Error: Can't Write to Open File {filename_out}")
        sys.exit()
    try:
        ws = wb["Data Summary"]
    except KeyError:
        ws = wb.create_sheet("Data Summary", 0)

    # define table titles and append to output workbook sheet "Data Summary"
    titles = [
        "Date",
        "Unit",
        "Battery",
        "Disposable",
        "Trial"
        "Input Temp Target (°C)",
        "Flowrate (mL/min)",
        "Input Temp Mean (°C)",
        "Steady-State Output Temp (°C)",
        "Reservoir Temp Mean (°C)",
        "Startup Time",
        "Peak Temp (°C)",
        "Test Time > 36°C",
        "Fluid Infused (mL)",
        "Battery Time",
        "ΔT x Time x Flowrate",
        "Filename",
        "Comment",
    ]
    if ws["A1"].value == None:
        ws.append(titles)
    # next line needed because sometimes append() skips over the first empty row
    if ws["A1"].value == None:
        ws.delete_rows(1)

    # define excel data table format and name
    table = Table(displayName="Summary", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    try:
        ws.add_table(table)
    except ValueError:
        pass

    # main loop to calculate and append data
    for filename in filenames:
        # read excel file
        try:
            file = pd.read_excel(filename, skiprows=6)
        except PermissionError:
            show_error(f"Error: Can't Write to Open File {filename}")
            continue
        try:
            input = file["Input (°C)"]
            output = file["Output (°C)"]
        except KeyError:
            show_error(f"Error: {filename} Invalid Data Structure")
            continue
        # reservoir data column is optional and will be skipped if not present
        try:
            reservoir = file["Reservoir (°C)"]
        except KeyError:
            pass

        # format time data
        time = file["Time"]
        file["Time"] = time - pd.Timestamp(time[0])

        # calculate data
        date = file["Date"].iloc[0]
        (
            flowrate,
            input_target,
            battery,
            disposable,
            unit,
            trial,
        ) = extract_filename_data(filename)
        if not flowrate:
            show_error(f"Error: {filename} Invalid Filename")
            continue
        peak_temp = output.max()
        if peak_temp < 36:
            battery_time = 0
            startup_time = "N/A"
            test_time = 0
            fluid_infused = 0
        else:
            battery_time = calculate_battery_time(file)
            startup_time, file_edit = strip_startup(file)
            test_time = calculate_test_time(file_edit)
            try:
                fluid_infused = round(flowrate * test_time.seconds / 60, 1)
            except AttributeError:
                fluid_infused = "error"
        input_mean = round(input.mean(), 2)
        output_mean = round(calculate_output_mean(file), 2)

        # reservoir data column is optional and will be skipped if not present
        try:
            reservoir_mean = round(reservoir.mean(), 2)
        except UnboundLocalError:
            reservoir_mean = "N/A"

        # calculates the product of delta temp, time and flowrate
        # this value is for reference only
        # this value should be similar for all tests that pass
        try:
            temp_time_flowrate = round(
                (output_mean - input_mean) * (test_time.seconds / 60 * flowrate / 1000),
                2,
            )
        except AttributeError:
            temp_time_flowrate = "error"

        # define list of data
        list = [
            date,
            unit,
            battery,
            disposable,
            trial,
            input_target,
            flowrate,
            input_mean,
            output_mean,
            reservoir_mean,
            startup_time,
            peak_temp,
            test_time,
            fluid_infused,
            battery_time,
            temp_time_flowrate,
            filename,
        ]
        # check if input file has already been edited
        # create new sheet "Data Summary" with calculated values
        wb_input = load_workbook(filename)
        if "Data Summary" in wb_input.sheetnames:
            del wb_input["Data Summary"]
        ws2 = wb_input.create_sheet("Data Summary")
        ws2.append(titles)
        ws2.append(list)
        try:
            wb_input.save(filename)
        except PermissionError:
            show_error(f"Error: Can't Write to Open File {filename}")
            continue

        # append data to "Data Summary Table" sheet in output file
        ws.append(list)
        ws.tables["Summary"].ref = ws.dimensions
        try:
            wb.save(filename_out)
        except PermissionError:
            show_error(f"Error: Can't Write to Open File {filename_out}")
            continue


# calculate the time the device delivered fluid above specfication (36°C)
# calculation starts from t = 0
# calculation ends once the output temp is below 36°C for 10 sec.
def calculate_test_time(file):
    timestep = file["Time"].iloc[1] - file["Time"].iloc[0]
    x = 10 / timestep.seconds
    count = 0
    for i in range(len(file)):
        if file["Output (°C)"].iloc[i] < 36:
            count += 1
            if count >= x:
                return file["Time"].iloc[i - 10]


# extract info from filename
# returns flowrate, input target temp, battery id#, disposable id#, unit id#, trial#
def extract_filename_data(filename):
    matches = re.search(
        r"([0-9]+)\w* ([0-9\.]+)\w* (?:batt)?_?(\w+)"
        r"(?: (?:disp)_?(\w+))?(?: (?:unit)_?(\w+))?(?: (?:trial)_?(\w+))?",
        filename,
        re.I,
    )
    if not matches:
        show_error(f"Error: {filename} Invalid Filename")
        return ("", "", "", "", "", "")
    # Define default for optional values.
    disp = "N/A"
    unit = "N/A"
    trial = "N/A"
    # Set optional values if they exist in filename
    if matches.group(4):
        disp = matches.group(4)
    if matches.group(5):
        unit = matches.group(5)
    if matches.group(6):
        trial = matches.group(6)
    return (
        int(matches.group(1)), # Flowrate
        int(matches.group(2)), # Input target temp
        matches.group(3), # Battery ID#
        disp,
        unit,
        trial,
    )


# returns how long until the battery ran out of charge
# calculation ends once the temperature drops below 30°C
# this calculation may need further refinement
def calculate_battery_time(file):
    file = file[(file["Time"] >= pd.Timedelta(5, "m")) & (file["Output (°C)"] <= 30)]
    try:
        return file["Time"].iloc[0]
    except IndexError:
        return "error"


# returns how long the device took to initially reach 36°C
# returns edited file that strips startup time
def strip_startup(file):
    file_temp = file[file["Output (°C)"] >= 36]
    startup_time = file_temp["Time"].iloc[0]
    file_edit = file[file["Time"] >= startup_time]
    return startup_time, file_edit


# returns output temperature during steady-state operation
# the data is assumed to be steady-stae between minutes 5 and 12
def calculate_output_mean(file):
    file_temp = file[
        (file["Time"] >= pd.Timedelta(5, unit="m"))
        & (file["Time"] <= pd.Timedelta(12, unit="m"))
    ]
    return file_temp["Output (°C)"].mean()


# shows error message box
def show_error(str):
    messagebox.showerror("error", str)


# apply wrap text format to cells
# not currently used
def styled_cells(data, ws):
    for c in data:
        c = Cell(ws, column="A", row=1, value=c)
        c.alignment = Alignment(wrap_text=True)
        yield c


if __name__ == "__main__":
    main()
