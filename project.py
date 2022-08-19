import pandas as pd
import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import filedialog, messagebox


def main():

    # select input file(s)
    filetypes = [("excel files", ".xlsx"), ("all files", ".*")]
    filenames = ""
    filenames = filedialog.askopenfilenames(
        initialdir=os.getcwd(), title="Select file input file(s)", filetypes=filetypes
    )
    if not filenames:
        show_error("Error: Select File(s)")
        sys.exit()

    # select file to append data
    # titles will be added to sheet "Data Summary"
    # if "Data Summary" sheet does not exit, one will be created
    filename_out = filedialog.askopenfilename(
        initialdir=os.getcwd(), title="Select output file", filetypes=filetypes
    )
    if not filename_out:
        show_error("Error: Select File")
        sys.exit()
    try:
        wb = load_workbook(filename_out)
    except PermissionError:
        show_error("Error: Can't Write to Open File")
        sys.exit()
    try:
        ws = wb["Data Summary"]
    except KeyError:
        ws = wb.create_sheet("Data Summary", 0)

    # define table titles and append to output workbook
    titles = [
        "Date",
        "Disposable",
        "Battery",
        "Input Temp Target (°C)",
        "Flowrate (mL/min)",
        "Input Temp Mean (°C)",
        "Steady-State Output Temp (°C)",
        "Reservoir Temp Mean (°C)",
        "Startup Time",
        "Peak Temp (°C)",
        "Test Time > 36°C",
        "Fluid Infused (mL)",
        "Battery Time"
        "ΔT x Time x Flowrate",
        "Comment",
    ]
    if ws["A1"].value == None:
        ws.append(titles)
    # next line needed because sometimes append() skips over the first empty row
    if ws["A1"].value == None:
        ws.delete_rows(1)

    # define excel data table format and name
    table = Table(displayName="Summary", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    try:
        ws.add_table(table)
    except ValueError:
        pass

    # main loop to calculate and append data
    for i in range(len(filenames)):
        # read excel file
        file = pd.read_excel(filenames[i], skiprows=6, usecols="A:E")
        try:
            input = file["Input (°C)"]
            output = file["Output (°C)"]
            # reservoir data column is optional and will be skipped if not present
            if not file["Reservoir (°C)"].empty:
                reservoir = file["Reservoir (°C)"]
        except KeyError:
            show_error(f"Error: {filenames[i]} Invalid Data Structure")
            sys.exit()

        # format time data
        time = file["Time"]
        file["Time"] = time - pd.Timestamp(time[0])

        # calculate data
        date = file["Date"].iloc[0]
        flowrate, input_target, battery, disposable = extract_filename_data(
            filenames[i]
        )
        peak_temp = output.max()
        if peak_temp < 36:
            battery_time = 0
            startup_time = "N/A"
            test_time = 0
            fluid_infused = 0
        else:
            battery_time = calculate_battery_time(file)
            startup_time, file_edit = strip_startup(file)
            test_time = calculate_test_time(file_edit)
            fluid_infused = flowrate * test_time.seconds / 60
        input_mean = round(input.mean(), 2)
        output_mean = round(calculate_output_mean(file), 2)

        # reservoir data column is optional and will be skipped if not present
        if not reservoir.empty:
            reservoir_mean = round(reservoir.mean(), 2)

        # calculates the product of delta temp, time and flowrate
        # this value is for reference only
        # this value should be similar for all tests that pass
        temp_time_flowrate = calculate_temp_time_flowrate(
            input_mean, output_mean, test_time.seconds, flowrate
        )

        # define list of data
        list = [
            date,
            disposable,
            battery,
            input_target,
            flowrate,
            input_mean,
            output_mean,
            reservoir_mean,
            startup_time,
            peak_temp,
            test_time,
            fluid_infused,
            battery_time,
            temp_time_flowrate,
        ]
        # check if input file has already been edited
        # create new sheet "Data Summary" with calculated values
        wb_input = load_workbook(filenames[i])
        if "Data Summary" in wb_input.sheetnames:
            show_error(f"Error: {filenames[i]} has already been edited")
            continue
        ws2 = wb_input.create_sheet("Data Summary")
        ws2.append(titles)
        ws2.append(list)
        try:
            wb_input.save(filenames[i])
        except PermissionError:
            show_error("Error: Can't Write to Open File")
            sys.exit()

        # append data to "Data Summary Table" sheet in output file
        ws.append(list)
        ws.tables["Summary"].ref = ws.dimensions
        try:
            wb.save(filename_out)
        except PermissionError:
            show_error("Error: Can't Write to Open File")
            sys.exit()


# returns the time the device delivered fluid above specfication (36°C)
# calculation starts from t = 0
# calculation ends once the output temp is below 36°C for 10 sec.
def calculate_test_time(file):
    timestep = file["Time"].iloc[1] - file["Time"].iloc[0]
    x = 10 / timestep.seconds
    count = 0
    for i in range(len(file)):
        if file["Output (°C)"].iloc[i] < 36:
            count += 1
            if count >= x:
                return file["Time"].iloc[i - 10]


# extract info from filename
# returns flowrate, input target temp, battery id#, disposable id#
def extract_filename_data(filename):
    matches = re.search(
        r"([0-9]+)\w* ([0-9]+)C? (\w+) (?:disp)?(\w+)\.xls", filename, re.I
    )
    if not matches:
        show_error(f"Error: {filename} Invalid Filename")
        sys.exit()
    return (
        int(matches.group(1)),
        int(matches.group(2)),
        matches.group(3),
        matches.group(4),
    )


# returns how long until the battery ran out of charge
# calculation ends once the temperature drops below 30°C
# this calculation may need further refinement
def calculate_battery_time(file):
    file = file[(file["Time"] >= pd.Timedelta(5, "m")) & (file["Output (°C)"] <= 30)]
    return file["Time"].iloc[0]


# returns how long the device took to initially reach 36°C
# returns edited file that strips startup time
def strip_startup(file):
    file_temp = file[file["Output (°C)"] >= 36]
    startup_time = file_temp["Time"].iloc[0]
    file_edit = file[file["Time"] >= startup_time]
    return startup_time, file_edit


# returns output temperature during steady-state operation
# the data is assumed to be steady-stae between minutes 5 and 12
def calculate_output_mean(file):
    file_temp = file[
        (file["Time"] >= pd.Timedelta(5, unit="m"))
        & (file["Time"] <= pd.Timedelta(12, unit="m"))
    ]
    return file_temp["Output (°C)"].mean()


# shows error message box
def show_error(str):
    messagebox.showerror("error", str)


def calculate_temp_time_flowrate(t0, t1, time, flowrate):
    return round((t1 - t0) * (time / 60 * flowrate / 1000), 2)


if __name__ == "__main__":
    main()
